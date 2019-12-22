from lxml import html,etree
import requests
from fake_useragent import UserAgent
import openpyxl
import sys
ua = UserAgent()
headers = {
     'User-Agent':ua.random
}
workbook = openpyxl.load_workbook('part2.xlsx')
sheet = workbook.worksheets[2]
row_count = sheet.max_row 
start_row = int(sys.argv[1])
end_row = int(sys.argv[2])
for i in range(start_row,end_row+1):
    resp = requests.get(url=sheet.cell(row=i,column=4).value,headers=headers)
    print("Row Number : ",i,"--->",resp.status_code)
    tree = html.fromstring(html=resp.content)
    description = tree.xpath("//div[@class='midd_sec']/p")[1:7] + tree.xpath("//div/table")
    sheet.cell(row=i,column=5).value = ''.join(list((html.tostring(descp, pretty_print=True).decode() for descp in description)))
workbook.save('part2.xlsx')