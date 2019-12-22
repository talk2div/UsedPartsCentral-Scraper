from lxml import html,etree
import requests
from fake_useragent import UserAgent
import openpyxl
ua = UserAgent()
headers = {
     'User-Agent':ua.random
}
workbook = openpyxl.load_workbook('part2.xlsx')
sheet = workbook.worksheets[1]
row_count = sheet.max_row
def geturl(imglst):
    try:
        return imglst[0]
    except IndexError:
        return ''
for i in range(2,row_count+1):
    resp = requests.get(url=sheet.cell(row=i,column=3).value,headers=headers)
    print("Row Number : ",i,"--->",resp.status_code)
    tree = html.fromstring(html=resp.content)
    description = tree.xpath("//div[@class='midd_sec']/p")[-3:-1]
    imgUrl = geturl(tree.xpath("//div[@class='fea_img']/img/@src"))
    sheet.cell(row=i,column=4).value = ''.join(list((html.tostring(descp, pretty_print=True).decode() for descp in description)))
    sheet.cell(row=i,column=5).value = imgUrl
workbook.save('part2.xlsx')