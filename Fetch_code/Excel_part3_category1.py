from lxml import html,etree
import requests
from fake_useragent import UserAgent
import openpyxl
ua = UserAgent()
headers = {
     'User-Agent':ua.random
}
workbook = openpyxl.load_workbook('part3.xlsx')
sheet = workbook.worksheets[0]
row_count = sheet.max_row
for i in range(2,row_count+1):
    resp = requests.get(url=sheet.cell(row=i,column=2).value,headers=headers)
    print("Row Number : ",i,"--->",resp.status_code)
    tree = html.fromstring(html=resp.content)
    description = tree.xpath('//div/img[@class="aligncenter"]') + tree.xpath("//div[@class='midd_sec']/p")[1:12]
    sheet.cell(row=i,column=3).value = ''.join(list((html.tostring(descp, pretty_print=True).decode() for descp in description)))
workbook.save('part3.xlsx')